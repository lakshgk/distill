"""
Tests for distill.parsers.xlsx — XlsxParser, XlsLegacyParser, and helpers.

Fixtures are built programmatically via openpyxl so no binary files are
checked into the repository.
"""

from __future__ import annotations

import csv
import io
import tempfile
import zipfile
from pathlib import Path

import pytest

from distill.ir import Document, Section, Table
from distill.parsers.base import ParseError
from distill.parsers.xlsx import (
    XlsLegacyParser,
    XlsxParser,
    _check_input_size,
    _check_zip_bomb,
    _expand_merged_cells,
    _rightmost_non_empty,
)


# ── Fixture helpers ──────────────────────────────────────────────────────────

def _make_xlsx(
    *,
    sheets: dict[str, list[list]] | None = None,
    title: str = "",
    author: str = "",
    subject: str = "",
    description: str = "",
    keywords: str = "",
) -> bytes:
    """Build a minimal .xlsx in memory and return its raw bytes."""
    import openpyxl

    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in (sheets or {"Sheet1": [["A", "B"], [1, 2]]}).items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)

    wb.properties.title       = title
    wb.properties.creator     = author
    wb.properties.subject     = subject
    wb.properties.description = description
    wb.properties.keywords    = keywords

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xlsx_with_merged(merge_range: str = "A1:B1") -> bytes:
    """Build a .xlsx where the given cell range is merged."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws.append(["Category", "", "Value"])
    ws.merge_cells(merge_range)
    ws.append(["Alpha", "Sub", 10])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xlsx_with_trailing_empty_cols() -> bytes:
    """Build a .xlsx with trailing empty columns on every row."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trim"
    ws.append(["Name", "Score", None, None, None])
    ws.append(["Alice", 95, None, None, None])
    ws.append(["Bob", 88, None, None, None])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv(rows: list[list], tmp_path: Path) -> Path:
    p = tmp_path / "data.csv"
    with open(p, "w", newline="") as f:
        csv.writer(f).writerows(rows)
    return p


def _cell_texts(table: Table) -> list[str]:
    return [
        run.text
        for row in table.rows
        for cell in row.cells
        for run in cell.content
    ]


def _tables(doc: Document) -> list[Table]:
    return [
        block
        for section in doc.sections
        for block in section.blocks
        if isinstance(block, Table)
    ]


# ── Parser availability ───────────────────────────────────────────────────────

class TestParserAvailability:
    def test_xlsx_is_available(self):
        assert XlsxParser.is_available()

    def test_xlsx_extensions(self):
        assert ".xlsx" in XlsxParser.extensions
        assert ".csv" in XlsxParser.extensions

    def test_xlsx_missing_requires_empty(self):
        assert XlsxParser.missing_requires() == []

    def test_xls_legacy_is_available(self):
        assert XlsLegacyParser.is_available()

    def test_xls_legacy_extension(self):
        assert ".xls" in XlsLegacyParser.extensions


# ── Basic parsing ─────────────────────────────────────────────────────────────

class TestBasicParsing:
    def test_returns_document(self):
        data   = _make_xlsx()
        result = XlsxParser().parse(data)
        assert isinstance(result, Document)

    def test_one_section_per_sheet(self):
        data = _make_xlsx(sheets={
            "Alpha": [["X"], [1]],
            "Beta":  [["Y"], [2]],
        })
        doc = XlsxParser().parse(data)
        headings = [r.text for s in doc.sections for r in (s.heading or [])]
        assert "Alpha" in headings
        assert "Beta"  in headings

    def test_section_level_is_2(self):
        data = _make_xlsx()
        doc  = XlsxParser().parse(data)
        assert all(s.level == 2 for s in doc.sections)

    def test_table_in_section(self):
        data = _make_xlsx(sheets={"Data": [["Col1", "Col2"], ["a", "b"]]})
        doc  = XlsxParser().parse(data)
        tbls = _tables(doc)
        assert len(tbls) == 1

    def test_first_row_is_header(self):
        data = _make_xlsx(sheets={"S": [["Name", "Age"], ["Alice", 30]]})
        doc  = XlsxParser().parse(data)
        header_row = _tables(doc)[0].rows[0]
        assert all(c.is_header for c in header_row.cells)

    def test_data_rows_not_header(self):
        data = _make_xlsx(sheets={"S": [["Name", "Age"], ["Alice", 30]]})
        doc  = XlsxParser().parse(data)
        data_row = _tables(doc)[0].rows[1]
        assert all(not c.is_header for c in data_row.cells)

    def test_cell_values_extracted(self):
        data = _make_xlsx(sheets={"S": [["Region", "Sales"], ["North", 500]]})
        doc  = XlsxParser().parse(data)
        texts = _cell_texts(_tables(doc)[0])
        assert "Region" in texts
        assert "North"  in texts
        assert "500"    in texts

    def test_accepts_path(self, tmp_path):
        p = tmp_path / "test.xlsx"
        p.write_bytes(_make_xlsx())
        doc = XlsxParser().parse(str(p))
        assert isinstance(doc, Document)

    def test_accepts_path_object(self, tmp_path):
        p = tmp_path / "test.xlsx"
        p.write_bytes(_make_xlsx())
        doc = XlsxParser().parse(p)
        assert isinstance(doc, Document)

    def test_multiple_sheets_all_extracted(self):
        data = _make_xlsx(sheets={
            "Sheet1": [["A"], [1]],
            "Sheet2": [["B"], [2]],
            "Sheet3": [["C"], [3]],
        })
        doc = XlsxParser().parse(data)
        assert len(doc.sections) == 3


# ── Metadata ──────────────────────────────────────────────────────────────────

class TestMetadata:
    def test_source_format(self):
        doc = XlsxParser().parse(_make_xlsx())
        assert doc.metadata.source_format == "xlsx"

    def test_sheet_count(self):
        data = _make_xlsx(sheets={"A": [["x"], [1]], "B": [["y"], [2]]})
        doc  = XlsxParser().parse(data)
        assert doc.metadata.sheet_count == 2

    def test_title(self, tmp_path):
        p = tmp_path / "t.xlsx"
        p.write_bytes(_make_xlsx(title="My Workbook"))
        doc = XlsxParser().parse(p)
        assert doc.metadata.title == "My Workbook"

    def test_author(self, tmp_path):
        p = tmp_path / "t.xlsx"
        p.write_bytes(_make_xlsx(author="Jane Doe"))
        doc = XlsxParser().parse(p)
        assert doc.metadata.author == "Jane Doe"

    def test_subject(self, tmp_path):
        p = tmp_path / "t.xlsx"
        p.write_bytes(_make_xlsx(subject="Q4 Finance"))
        doc = XlsxParser().parse(p)
        assert doc.metadata.subject == "Q4 Finance"

    def test_description(self, tmp_path):
        p = tmp_path / "t.xlsx"
        p.write_bytes(_make_xlsx(description="Annual report data"))
        doc = XlsxParser().parse(p)
        assert doc.metadata.description == "Annual report data"

    def test_keywords_comma(self, tmp_path):
        p = tmp_path / "t.xlsx"
        p.write_bytes(_make_xlsx(keywords="finance, q4, 2025"))
        doc = XlsxParser().parse(p)
        assert "finance" in doc.metadata.keywords
        assert "q4"      in doc.metadata.keywords

    def test_keywords_semicolon(self, tmp_path):
        p = tmp_path / "t.xlsx"
        p.write_bytes(_make_xlsx(keywords="alpha;beta;gamma"))
        doc = XlsxParser().parse(p)
        assert doc.metadata.keywords == ["alpha", "beta", "gamma"]


# ── Merged cells ──────────────────────────────────────────────────────────────

class TestMergedCells:
    def test_merged_value_repeated(self):
        data = _make_xlsx_with_merged("A1:B1")
        doc  = XlsxParser().parse(data)
        tbls = _tables(doc)
        assert len(tbls) == 1
        header_texts = [run.text for c in tbls[0].rows[0].cells for run in c.content]
        assert header_texts.count("Category") == 2

    def test_helper_expand_merged_cells(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Merged", None, "C"])
        ws.merge_cells("A1:B1")
        result = _expand_merged_cells(ws)
        assert result[(1, 1)] == "Merged"
        assert result[(1, 2)] == "Merged"
        assert (1, 3) not in result

    def test_row_span_repeats_anchor_across_rows(self):
        """A cell merged across 3 rows repeats the anchor value in all 3 output rows."""
        fixture = Path(__file__).parent / "fixtures" / "merged_cells.xlsx"
        doc = XlsxParser().parse(fixture)
        tbls = _tables(doc)
        assert len(tbls) == 1
        # Rows 1-3 (data rows after the header) should all have "North" in column 0
        data_rows = tbls[0].rows[1:]  # skip header
        col0_values = [row.cells[0].content[0].text for row in data_rows]
        assert col0_values == ["North", "North", "North"]

    def test_col_span_repeats_anchor_across_columns(self):
        """A cell merged across 2 columns repeats the anchor value in both output columns."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ColMerge"
        ws["A1"] = "Header"
        ws["B1"] = "H2"
        ws["C1"] = "H3"
        ws["A2"] = "Span"
        ws["C2"] = "Solo"
        ws.merge_cells("A2:B2")
        ws.append(["x", "y", "z"])

        buf = io.BytesIO()
        wb.save(buf)
        doc = XlsxParser().parse(buf.getvalue())
        tbls = _tables(doc)
        row1_texts = [c.content[0].text for c in tbls[0].rows[1].cells]
        assert row1_texts[0] == "Span"
        assert row1_texts[1] == "Span"
        assert row1_texts[2] == "Solo"

    def test_no_merged_cells_unchanged(self):
        """A workbook with no merged cells produces correct output (regression guard)."""
        fixture = Path(__file__).parent / "fixtures" / "merged_cells.xlsx"
        # Build a plain workbook with identical data but no merges
        plain = _make_xlsx(sheets={"Sales": [
            ["Region", "Q1", "Q2"],
            ["North", "12", "14"],
            ["North", "8", "9"],
            ["North", "5", "6"],
        ]})
        doc_plain = XlsxParser().parse(plain)
        doc_fixture = XlsxParser().parse(fixture)
        # Both should produce tables with the same cell texts
        plain_texts = _cell_texts(_tables(doc_plain)[0])
        fixture_texts = _cell_texts(_tables(doc_fixture)[0])
        assert plain_texts == fixture_texts

    def test_none_anchor_yields_empty_string(self):
        """A merged cell whose anchor value is None outputs an empty string, not 'None'."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NoneAnchor"
        ws["A1"] = "H1"
        ws["B1"] = "H2"
        # A2 is left as None, then merged across A2:A3
        ws["B2"] = "val1"
        ws["B3"] = "val2"
        ws.merge_cells("A2:A3")

        buf = io.BytesIO()
        wb.save(buf)
        doc = XlsxParser().parse(buf.getvalue())
        tbls = _tables(doc)
        # Both data rows should have empty string in column 0
        for row in tbls[0].rows[1:]:
            text = row.cells[0].content[0].text
            assert text == "", f"Expected empty string, got {text!r}"

    def test_merge_fix_applies_to_all_sheets(self):
        """The merge fix applies to every worksheet, not just the first."""
        import openpyxl
        wb = openpyxl.Workbook()
        # Sheet 1 — merge in column A rows 2-3
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "Key"
        ws1["B1"] = "Val"
        ws1["A2"] = "Alpha"
        ws1["B2"] = "1"
        ws1["B3"] = "2"
        ws1.merge_cells("A2:A3")

        # Sheet 2 — merge in column A rows 2-3
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "Key"
        ws2["B1"] = "Val"
        ws2["A2"] = "Beta"
        ws2["B2"] = "3"
        ws2["B3"] = "4"
        ws2.merge_cells("A2:A3")

        buf = io.BytesIO()
        wb.save(buf)
        doc = XlsxParser().parse(buf.getvalue())

        for section in doc.sections:
            tbl = [b for b in section.blocks if isinstance(b, Table)][0]
            col0 = [row.cells[0].content[0].text for row in tbl.rows[1:]]
            expected = section.heading[0].text.replace("Sheet1", "Alpha").replace("Sheet2", "Beta")
            assert col0 == [expected, expected], (
                f"Sheet {section.heading[0].text}: expected [{expected!r}, {expected!r}], got {col0}"
            )


# ── Word count ────────────────────────────────────────────────────────────────

class TestWordCount:
    def test_word_count_populated(self):
        fixture = Path(__file__).parent / "fixtures" / "merged_cells.xlsx"
        doc = XlsxParser().parse(fixture)
        assert doc.metadata.word_count is not None
        assert doc.metadata.word_count > 0


# ── Empty column trimming ─────────────────────────────────────────────────────

class TestEmptyColumnTrim:
    def test_trailing_empty_cols_removed(self):
        data = _make_xlsx_with_trailing_empty_cols()
        doc  = XlsxParser().parse(data)
        tbls = _tables(doc)
        assert len(tbls) == 1
        col_count = len(tbls[0].rows[0].cells)
        assert col_count == 2

    def test_rightmost_non_empty_helper(self):
        rows = [["A", "B", "", ""], ["1", "2", "", ""]]
        assert _rightmost_non_empty(rows) == 2

    def test_rightmost_non_empty_all_empty(self):
        rows = [["", ""], ["", ""]]
        assert _rightmost_non_empty(rows) == 0

    def test_rightmost_non_empty_last_col_populated(self):
        rows = [["", "", "C"], ["", "", "3"]]
        assert _rightmost_non_empty(rows) == 3


# ── Row truncation ────────────────────────────────────────────────────────────

class TestRowTruncation:
    def test_truncation_applied(self):
        from distill.parsers.base import ParseOptions
        rows = [["H"]] + [[str(i)] for i in range(200)]
        data = _make_xlsx(sheets={"Big": rows})
        opts = ParseOptions(max_table_rows=10)
        doc  = XlsxParser().parse(data, options=opts)
        tbls = _tables(doc)
        assert len(tbls[0].rows) == 10

    def test_truncation_warning_emitted(self):
        from distill.parsers.base import ParseOptions
        rows = [["H"]] + [[str(i)] for i in range(200)]
        data = _make_xlsx(sheets={"Big": rows})
        opts = ParseOptions(max_table_rows=10)
        doc  = XlsxParser().parse(data, options=opts)
        assert any("truncated" in w for w in doc.warnings)

    def test_no_truncation_when_unlimited(self):
        from distill.parsers.base import ParseOptions
        rows = [["H"]] + [[str(i)] for i in range(600)]
        data = _make_xlsx(sheets={"Big": rows})
        opts = ParseOptions(max_table_rows=0)
        doc  = XlsxParser().parse(data, options=opts)
        tbls = _tables(doc)
        assert len(tbls[0].rows) == 601


# ── CSV path ──────────────────────────────────────────────────────────────────

class TestCsvParsing:
    def test_csv_returns_document(self, tmp_path):
        p   = _make_csv([["Name", "Age"], ["Alice", "30"]], tmp_path)
        doc = XlsxParser().parse(p)
        assert isinstance(doc, Document)

    def test_csv_source_format(self, tmp_path):
        p   = _make_csv([["X"], ["1"]], tmp_path)
        doc = XlsxParser().parse(p)
        assert doc.metadata.source_format == "csv"

    def test_csv_values_extracted(self, tmp_path):
        p    = _make_csv([["City", "Pop"], ["Oslo", "700000"]], tmp_path)
        doc  = XlsxParser().parse(p)
        tbls = _tables(doc)
        texts = _cell_texts(tbls[0])
        assert "City"   in texts
        assert "Oslo"   in texts
        assert "700000" in texts

    def test_csv_first_row_header(self, tmp_path):
        p   = _make_csv([["H1", "H2"], ["v1", "v2"]], tmp_path)
        doc = XlsxParser().parse(p)
        tbls = _tables(doc)
        assert all(c.is_header for c in tbls[0].rows[0].cells)

    def test_csv_trailing_empty_cols_trimmed(self, tmp_path):
        p   = _make_csv([["A", "B", "", ""], ["1", "2", "", ""]], tmp_path)
        doc = XlsxParser().parse(p)
        tbls = _tables(doc)
        assert len(tbls[0].rows[0].cells) == 2


# ── Security ──────────────────────────────────────────────────────────────────

class TestSecurity:
    def test_input_size_bytes_rejected(self):
        oversized = b"x" * (55 * 1024 * 1024)
        with pytest.raises(ParseError, match="50 MB"):
            _check_input_size(oversized, 50 * 1024 * 1024)

    def test_input_size_path_rejected(self, tmp_path):
        big = tmp_path / "big.xlsx"
        big.write_bytes(b"x" * (55 * 1024 * 1024))
        with pytest.raises(ParseError, match="50 MB"):
            _check_input_size(str(big), 50 * 1024 * 1024)

    def test_input_size_ok(self):
        _check_input_size(b"x" * 100, 50 * 1024 * 1024)  # must not raise

    def test_zip_bomb_detected(self):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("big.bin", "A" * (501 * 1024 * 1024))
        with pytest.raises(ParseError, match="500 MB"):
            _check_zip_bomb(buf.getvalue(), 500 * 1024 * 1024)

    def test_bad_zip_raises(self):
        with pytest.raises(ParseError, match="not a valid XLSX"):
            _check_zip_bomb(b"not a zip file", 500 * 1024 * 1024)

    def test_parser_rejects_oversized(self):
        oversized = b"x" * (55 * 1024 * 1024)
        with pytest.raises(ParseError, match="50 MB"):
            XlsxParser().parse(oversized)

    def test_custom_size_limit(self):
        from distill.parsers.base import ParseOptions
        data = b"x" * (15 * 1024 * 1024)
        opts = ParseOptions(extra={"max_file_size": 10 * 1024 * 1024})
        with pytest.raises(ParseError, match="10 MB"):
            XlsxParser().parse(data, options=opts)


# ── XlsLegacyParser ──────────────────────────────────────────────────────────

class TestXlsLegacyParser:
    def test_raises_parse_error(self):
        # Wired to LibreOffice — will fail due to binary not found or conversion error
        with pytest.raises(ParseError):
            XlsLegacyParser().parse(b"garbage")

    def test_error_mentions_libreoffice(self):
        with pytest.raises(ParseError, match="LibreOffice"):
            XlsLegacyParser().parse(b"garbage")

    def test_extension_registered(self):
        assert ".xls" in XlsLegacyParser.extensions


# ── Render integration ─────────────────────────────────────────────────────────

class TestRenderIntegration:
    def test_renders_to_markdown(self):
        data = _make_xlsx(sheets={"Data": [["Item", "Qty"], ["Apples", 5]]})
        doc  = XlsxParser().parse(data)
        md   = doc.render(front_matter=False)
        assert isinstance(md, str)
        assert "Item" in md or "Apples" in md

    def test_front_matter_has_source_format(self):
        data = _make_xlsx()
        doc  = XlsxParser().parse(data)
        md   = doc.render(front_matter=True)
        assert "---" in md
        assert "xlsx" in md

    def test_no_front_matter_when_suppressed(self):
        data = _make_xlsx()
        doc  = XlsxParser().parse(data)
        md   = doc.render(front_matter=False)
        # The YAML front-matter block starts with "---\n"; GFM table separators
        # also contain "---" but not at position 0 of the output.
        assert not md.startswith("---")

    def test_sheet_name_as_heading(self):
        data = _make_xlsx(sheets={"MySales": [["A"], [1]]})
        doc  = XlsxParser().parse(data)
        md   = doc.render(front_matter=False)
        assert "MySales" in md

    def test_table_pipe_syntax_in_output(self):
        data = _make_xlsx(sheets={"T": [["Col1", "Col2"], ["r1", "r2"]]})
        doc  = XlsxParser().parse(data)
        md   = doc.render(front_matter=False)
        assert "|" in md


# ── XLSM support ─────────────────────────────────────────────────────────────

class TestXlsmSupport:
    XLSM_FIXTURE = Path(__file__).parent / "fixtures" / "simple.xlsm"

    def test_registry_finds_xlsm(self):
        from distill.registry import registry
        entry = registry.find("test.xlsm")
        assert entry is not None

    def test_xlsm_routes_to_xlsx_parser(self):
        from distill.registry import registry
        xlsm = registry.find("test.xlsm")
        xlsx = registry.find("test.xlsx")
        assert xlsm == xlsx

    def test_xlsm_mime_type_accepted(self):
        assert "application/vnd.ms-excel.sheet.macroEnabled.12" in XlsxParser.mime_types

    def test_xlsm_produces_document_with_sections(self):
        from distill.parsers.base import ParseOptions
        from distill.warnings import WarningCollector
        opts = ParseOptions()
        opts.collector = WarningCollector()
        doc = XlsxParser().parse(self.XLSM_FIXTURE, options=opts)
        assert len(doc.sections) >= 1

    def test_xlsm_emits_macro_warning(self):
        from distill.parsers.base import ParseOptions
        from distill.warnings import WarningCollector
        opts = ParseOptions()
        opts.collector = WarningCollector()
        XlsxParser().parse(self.XLSM_FIXTURE, options=opts)
        warnings = opts.collector.to_dict()
        assert len(warnings) >= 1
        assert any("macro" in w["message"].lower() for w in warnings)

    def test_xlsx_does_not_emit_macro_warning(self):
        from distill.parsers.base import ParseOptions
        from distill.warnings import WarningCollector
        opts = ParseOptions()
        opts.collector = WarningCollector()
        data = _make_xlsx()
        XlsxParser().parse(data, options=opts)
        warnings = opts.collector.to_dict()
        assert not any("macro" in w.get("message", "").lower() for w in warnings)
