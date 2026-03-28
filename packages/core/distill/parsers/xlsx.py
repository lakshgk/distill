"""
distill.parsers.xlsx
~~~~~~~~~~~~~~~~~~~~
Parser for Microsoft Excel workbooks (.xlsx) and CSV files.

Primary path:   openpyxl (.xlsx structural read, data_only=True for computed values)
Lightweight:    csv module (for plain .csv files)
Legacy stub:    .xls raises ParseError with LibreOffice install hint (Phase 2)

Key design decisions:
- Each worksheet becomes a separate H2 Section
- Formulas: render cached computed value (data_only=True).  If the cache is
  empty (file never opened in Excel), a warning is emitted per sheet.
- Merged cells: the top-left value is repeated into subordinate cells so the
  table row is always fully populated
- Trailing empty columns: detected and trimmed from every sheet
- Row cap: ParseOptions.max_table_rows (default 500); 0 = unlimited
- Empty / chart-only sheets: skipped with a warning
- Security: 50 MB input size limit; 500 MB zip bomb limit (xlsx is a ZIP)
"""

from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path
from typing import Optional, Union

from distill.ir import (
    Document, DocumentMetadata, Section, Table,
    TableCell, TableRow, TextRun,
)
from distill.parsers.base import ParseError, ParseOptions, Parser
from distill.registry import registry


# ── Security constants ────────────────────────────────────────────────────────

_MAX_FILE_BYTES  = 50  * 1024 * 1024   # 50 MB
_MAX_UNZIP_BYTES = 500 * 1024 * 1024   # 500 MB


def _check_input_size(source: Union[str, Path, bytes], max_bytes: int) -> None:
    """Raise ParseError if the source exceeds max_bytes."""
    mb = max_bytes // (1024 * 1024)
    if isinstance(source, (str, Path)):
        size = Path(source).stat().st_size
        if size > max_bytes:
            raise ParseError(
                f"Input file exceeds the {mb} MB size limit "
                f"({size / (1024*1024):.1f} MB). "
                f"Increase the limit via options.extra['max_file_size']."
            )
    elif isinstance(source, bytes):
        if len(source) > max_bytes:
            raise ParseError(
                f"Input file exceeds the {mb} MB size limit "
                f"({len(source) / (1024*1024):.1f} MB). "
                f"Increase the limit via options.extra['max_file_size']."
            )


def _check_zip_bomb(source: Union[str, Path, bytes], max_unzip_bytes: int) -> None:
    """Raise ParseError if the ZIP uncompressed size exceeds max_unzip_bytes."""
    mb = max_unzip_bytes // (1024 * 1024)
    try:
        data = source if isinstance(source, bytes) else Path(source).read_bytes()
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            total = sum(info.file_size for info in zf.infolist())
        if total > max_unzip_bytes:
            raise ParseError(
                f"XLSX archive uncompressed size ({total // (1024*1024)} MB) "
                f"exceeds the {mb} MB safety limit. "
                f"Increase via options.extra['max_unzip_size']."
            )
    except ParseError:
        raise
    except zipfile.BadZipFile:
        raise ParseError("File is not a valid XLSX (ZIP) archive.")
    except Exception as e:
        raise ParseError(f"Could not inspect archive: {e}") from e


# ── Metadata helpers ──────────────────────────────────────────────────────────

def _extract_metadata(wb, path: Optional[Path], sheet_count: int) -> DocumentMetadata:
    """
    Pull core properties from the openpyxl workbook properties object.
    openpyxl exposes workbook.properties (a PackagePropertiesPart) with the
    same OOXML core-property fields as python-docx core_properties.
    """
    props = wb.properties

    def _safe(attr: str) -> Optional[str]:
        try:
            v = getattr(props, attr, None)
            return str(v).strip() if v is not None else None
        except Exception:
            return None

    def _iso(attr: str) -> Optional[str]:
        try:
            v = getattr(props, attr, None)
            return v.isoformat() if v is not None else None
        except Exception:
            return None

    kw_raw   = _safe("keywords") or ""
    keywords = [k.strip() for k in re.split(r"[,;]", kw_raw) if k.strip()] if kw_raw else []

    return DocumentMetadata(
        title         = _safe("title"),
        author        = _safe("creator"),
        subject       = _safe("subject"),
        description   = _safe("description"),
        keywords      = keywords,
        created_at    = _iso("created"),
        modified_at   = _iso("modified"),
        sheet_count   = sheet_count,
        source_format = "xlsx",
        source_path   = str(path) if path else None,
    )


# ── Column utilities ──────────────────────────────────────────────────────────

def _rightmost_non_empty(rows_data: list[list[str]]) -> int:
    """Return the column width (exclusive) of the rightmost non-empty cell."""
    max_col = 0
    for row in rows_data:
        for i, cell in enumerate(row):
            if cell.strip():
                max_col = max(max_col, i + 1)
    return max_col


def _expand_merged_cells(ws) -> dict[tuple[int, int], str]:
    """
    Build a (row, col) → value map for merged-cell regions.
    The top-left cell value is repeated into every subordinate position.
    Row and column indices are 1-based (matching openpyxl cell.row/cell.column).
    """
    merged_values: dict[tuple[int, int], str] = {}
    for merge_range in ws.merged_cells.ranges:
        top_left = ws.cell(row=merge_range.min_row, column=merge_range.min_col)
        value    = str(top_left.value) if top_left.value is not None else ""
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merged_values[(row, col)] = value
    return merged_values


def _has_formula_cache(ws) -> bool:
    """Return True if at least one formula cell has a cached computed value."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f" and cell.value is not None:
                return True
    return False


# ── Parser ────────────────────────────────────────────────────────────────────

@registry.register
class XlsxParser(Parser):
    """Parses .xlsx workbooks using openpyxl."""

    extensions = [".xlsx", ".csv"]
    mime_types = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "text/csv",
    ]
    requires          = ["openpyxl"]
    optional_requires = []

    def parse(
        self,
        source: Union[str, Path, bytes],
        options: Optional[ParseOptions] = None,
    ) -> Document:
        options = options or ParseOptions()

        if isinstance(source, (str, Path)):
            path = Path(source)
        else:
            path = None

        if path and path.suffix.lower() == ".csv":
            return self._parse_csv(path, options)

        max_file  = options.extra.get("max_file_size",  _MAX_FILE_BYTES)
        max_unzip = options.extra.get("max_unzip_size", _MAX_UNZIP_BYTES)
        _check_input_size(source, max_file)
        _check_zip_bomb(source, max_unzip)

        return self._parse_xlsx(path, source, options)

    def _parse_xlsx(
        self,
        path: Optional[Path],
        source: Union[str, Path, bytes],
        options: ParseOptions,
    ) -> Document:
        try:
            import openpyxl
        except ImportError as e:
            raise ParseError(f"Missing dependency: {e}") from e

        try:
            if path:
                wb = openpyxl.load_workbook(str(path), data_only=True)
            else:
                wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)  # type: ignore[arg-type]
        except Exception as e:
            raise ParseError(f"openpyxl failed to open workbook: {e}") from e

        sheet_count = len(wb.sheetnames)
        metadata    = _extract_metadata(wb, path, sheet_count)
        document    = Document(metadata=metadata)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            if ws.max_row is None or ws.max_row == 0:
                document.warnings.append(f"[xlsx] Skipped empty sheet: {sheet_name!r}")
                continue

            section = Section(
                heading=[TextRun(text=sheet_name)],
                level=2,
            )

            table = self._worksheet_to_table(ws, options, sheet_name, document)
            if table is not None:
                section.blocks.append(table)

            document.sections.append(section)

        return document

    def _worksheet_to_table(
        self,
        ws,
        options: ParseOptions,
        sheet_name: str,
        doc: Document,
    ) -> Optional[Table]:
        merged = _expand_merged_cells(ws)

        # Warn if formula cells have no cached values
        has_formulas = any(
            cell.data_type == "f"
            for row in ws.iter_rows()
            for cell in row
        )
        if has_formulas and not _has_formula_cache(ws):
            doc.warnings.append(
                f"[xlsx] Sheet {sheet_name!r}: formula cells have no cached values. "
                f"Open and re-save the workbook in Excel to populate computed values."
            )

        rows_data: list[list[str]] = []
        for row in ws.iter_rows():
            row_cells: list[str] = []
            for cell in row:
                value = merged.get((cell.row, cell.column))
                if value is None:
                    value = str(cell.value) if cell.value is not None else ""
                row_cells.append(value)
            rows_data.append(row_cells)

        if not rows_data:
            return None

        # Trim trailing empty columns
        width = _rightmost_non_empty(rows_data)
        if width == 0:
            doc.warnings.append(f"[xlsx] Skipped all-empty sheet: {sheet_name!r}")
            return None
        rows_data = [row[:width] for row in rows_data]

        total_rows = len(rows_data)
        truncated  = False
        if options.max_table_rows > 0 and total_rows > options.max_table_rows:
            rows_data = rows_data[:options.max_table_rows]
            truncated = True
            doc.warnings.append(
                f"[xlsx] Sheet {sheet_name!r}: truncated to {options.max_table_rows} "
                f"of {total_rows} rows."
            )

        def make_row(cells: list[str], is_header: bool = False) -> TableRow:
            return TableRow(cells=[
                TableCell(
                    content=[TextRun(text=cell)],
                    is_header=is_header,
                )
                for cell in cells
            ])

        table_rows = [make_row(rows_data[0], is_header=True)]
        for row in rows_data[1:]:
            table_rows.append(make_row(row))

        return Table(
            rows=table_rows,
            truncated=truncated,
            total_rows=total_rows if truncated else None,
        )

    def _parse_csv(self, path: Path, options: ParseOptions) -> Document:
        import csv

        metadata = DocumentMetadata(source_format="csv", source_path=str(path))
        document = Document(metadata=metadata)

        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                rows_data = list(csv.reader(f))
        except Exception as e:
            raise ParseError(f"Could not read CSV file: {e}") from e

        if not rows_data:
            return document

        total_rows = len(rows_data)
        truncated  = False
        if options.max_table_rows > 0 and total_rows > options.max_table_rows:
            rows_data = rows_data[:options.max_table_rows]
            truncated = True

        width = _rightmost_non_empty(rows_data)
        if width > 0:
            rows_data = [row[:width] for row in rows_data]

        def make_row(cells: list[str], is_header: bool = False) -> TableRow:
            return TableRow(cells=[
                TableCell(content=[TextRun(text=c)], is_header=is_header)
                for c in cells
            ])

        table_rows = [make_row(rows_data[0], is_header=True)]
        for row in rows_data[1:]:
            table_rows.append(make_row(row))

        table   = Table(rows=table_rows, truncated=truncated, total_rows=total_rows if truncated else None)
        section = Section(level=0, blocks=[table])
        document.sections.append(section)
        return document


@registry.register
class XlsLegacyParser(Parser):
    """
    Converts legacy .xls binary workbooks to .xlsx via LibreOffice headless,
    then delegates to XlsxParser for the actual content extraction.

    Requires LibreOffice to be installed and on PATH (or DISTILL_LIBREOFFICE
    environment variable set to the full binary path).
    """

    extensions            = [".xls"]
    mime_types            = ["application/vnd.ms-excel"]
    requires              = ["openpyxl"]
    requires_libreoffice  = True

    def parse(
        self,
        source: Union[str, Path, bytes],
        options: Optional[ParseOptions] = None,
    ) -> Document:
        import shutil
        from distill.parsers._libreoffice import convert_via_libreoffice

        options = options or ParseOptions()
        timeout = options.extra.get("libreoffice_timeout", 60)

        output_path = convert_via_libreoffice(source, "xlsx", timeout=timeout)
        try:
            doc = XlsxParser().parse(output_path, options)
            # Preserve the original source path/format in metadata
            doc.metadata.source_format = "xls"
            if not isinstance(source, bytes):
                doc.metadata.source_path = str(source)
            return doc
        finally:
            shutil.rmtree(output_path.parent, ignore_errors=True)
